#!/usr/bin/env python3
"""
price-book.xlsx  ->  workbook-snapshot.json

Read-only extraction of the estimating workbook for the app import pipeline.

WHAT IT DOES
  1. Copies the live workbook to a temp file. Everything after this touches the COPY.
     The live file is opened exactly once, by shutil.copy2, for reading.
  2. Validates the workbook's SHAPE against workbook-mapping.json. A missing tab, a
     renamed header, or a header at the wrong column letter aborts the run naming
     exactly what was expected and what was found. Nothing is guessed by position.
  3. Reads structure and inputs with openpyxl (formulas, not values).
  4. Recalculates the COPY in Excel to obtain the workbook's own computed totals,
     because the live file carries no cached formula values (see mapping file,
     knownStructuralFacts.noCachedFormulaValues). Excel opens the copy READ-ONLY and
     is closed with SaveChanges:=False. If Excel is unavailable the run still
     succeeds and emits `workbookComputed: null` -- the importer then refuses to
     claim parity rather than inventing a comparison.
  5. Verifies the live workbook's SHA-256 is unchanged, start to finish, and records
     both hashes in the snapshot.

WHY PYTHON: openpyxl and the Excel COM bridge are the toolchain the 02:00/03:03
scheduled price-book tasks already use against this same file. The import and parity
logic live in TypeScript with the rest of the app; this script is the read boundary.

USAGE
  python extract_workbook.py --workbook <path> --out <snapshot.json> [--no-recalc]

EXIT CODES
  0  snapshot written
  2  shape drift -- workbook does not match workbook-mapping.json
  3  workbook not found / not readable
  4  internal consistency failure (e.g. cost and sell formulas disagree)
"""

import argparse
import hashlib
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
from datetime import datetime, timezone

try:
    import openpyxl
    from openpyxl.utils import get_column_letter, column_index_from_string
except ImportError:
    print("FATAL: openpyxl is required. pip install openpyxl", file=sys.stderr)
    sys.exit(3)

HERE = os.path.dirname(os.path.abspath(__file__))
MAPPING_PATH = os.path.join(HERE, "workbook-mapping.json")


# ─── helpers ────────────────────────────────────────────────────────────────────

def norm(s):
    """Normalise header/label text for comparison: collapse whitespace, strip, casefold."""
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s)).strip().casefold()


def sha256(path):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


class ShapeDrift(Exception):
    """Raised when the workbook no longer matches the mapping contract."""


class Drifts:
    """Collects every shape problem so the operator sees all of them at once,
    not one per re-run."""

    def __init__(self):
        self.items = []

    def add(self, where, expected, found):
        self.items.append({"where": where, "expected": expected, "found": found})

    def raise_if_any(self):
        if not self.items:
            return
        lines = ["WORKBOOK SHAPE DRIFT -- import aborted. Nothing was written.", ""]
        for d in self.items:
            lines.append(f"  {d['where']}")
            lines.append(f"     expected: {d['expected']!r}")
            lines.append(f"     found:    {d['found']!r}")
        lines += [
            "",
            "This is the IMPORT MAPPING IMPACT case. Update",
            "  app/scripts/price-book/workbook-mapping.json",
            "to match the workbook's new shape, then re-run. Do not edit the workbook to",
            "match the mapping -- the workbook is the source of truth.",
        ]
        raise ShapeDrift("\n".join(lines))


# ─── shape validation ───────────────────────────────────────────────────────────

def validate_tab(wb, key, spec, mapping, drifts):
    sheet = spec["sheet"]
    if sheet not in wb.sheetnames:
        drifts.add(f"tab {key!r}", f"sheet named {sheet!r}", f"sheets present: {wb.sheetnames}")
        return None
    ws = wb[sheet]

    cols = spec.get("columns")
    if cols is None and "sameColumnsAs" in spec:
        cols = mapping["tabs"][spec["sameColumnsAs"]]["columns"]
    prefix_only = spec.get("headerPrefixOnly")
    if prefix_only is None and "sameColumnsAs" in spec:
        prefix_only = mapping["tabs"][spec["sameColumnsAs"]].get("headerPrefixOnly", {})
    prefix_only = prefix_only or {}

    hrow = mapping.get("headerRow", 1)
    for letter, expected in (cols or {}).items():
        idx = column_index_from_string(letter)
        found = ws.cell(hrow, idx).value
        if norm(found) != norm(expected):
            drifts.add(f"{sheet}!{letter}{hrow}", expected, found)

    # Columns whose headers carry long provenance sentences that are edited in place.
    # Matching the whole string would make every annotation a false shape drift, so
    # these are matched on a stable leading token instead.
    for letter, expected_prefix in prefix_only.items():
        idx = column_index_from_string(letter)
        found = ws.cell(hrow, idx).value
        if not norm(found).startswith(norm(expected_prefix)):
            drifts.add(f"{sheet}!{letter}{hrow} (prefix match)", expected_prefix + "...", found)

    return ws


def validate_rate_config(wb, mapping, drifts):
    spec = mapping["rateConfig"]
    sheet = spec["sheet"]
    if sheet not in wb.sheetnames:
        drifts.add("rateConfig", f"sheet named {sheet!r}", f"sheets present: {wb.sheetnames}")
        return None
    ws = wb[sheet]
    lcol = column_index_from_string(spec["labelColumn"])
    for key, cell in spec["cells"].items():
        found = ws.cell(cell["row"], lcol).value
        if norm(found) != norm(cell["label"]):
            drifts.add(
                f"{sheet}!{spec['labelColumn']}{cell['row']} (label for {key})",
                cell["label"],
                found,
            )
    return ws


# ─── extraction ─────────────────────────────────────────────────────────────────

def cell_or_none(ws, row, letter):
    v = ws.cell(row, column_index_from_string(letter)).value
    if isinstance(v, str) and v.strip() == "":
        return None
    return v


def as_number(v):
    """Return a float, or None. NEVER returns 0 for a blank -- a blank price is
    absent, not free (CLAUDE.md: never make up a number)."""
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("$", "").replace(",", "")
    if s == "":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_labor_unit_basis(raw):
    """Resolve Atomics!AA to a NECA unit letter and its divisor.

    The column holds a bare 'E' / 'C' / 'M' on most rows, and an annotated form where the
    04:00 run cited its source, e.g.
        'E  [NECA p.268 "Two Pole Circuit Breaker 100 Amp 0.94 1.18 1.41 E" — read verbatim]'
    so the letter is matched as a FIRST TOKEN, not as the whole string.

    Everything else — 'UNVERIFIED …', 'n/a …', blank — resolves to (None, None). That is the
    column header's own instruction, quoted verbatim there: "the app must BLOCK, not default
    to E." A wrong divisor here is a 100x labour error that would look like a plausible number,
    which is precisely the failure CLAUDE.md calls the most dangerous thing to hand over.
    """
    if raw is None:
        return None, None
    token = str(raw).strip().split()
    if not token:
        return None, None
    letter = token[0].strip().upper()
    divisors = {"E": 1.0, "C": 100.0, "M": 1000.0}
    if letter in divisors:
        return letter, divisors[letter]
    return None, None


def extract_atomics(ws, spec):
    cols = spec["columns"]
    out, row_to_id = [], {}
    for r in range(2, ws.max_row + 1):
        item_id = cell_or_none(ws, r, "A")
        if item_id is None:
            continue
        item_id = str(item_id).strip()
        row_to_id[r] = item_id
        basis_raw = cell_or_none(ws, r, "AA")
        basis_letter, basis_divisor = parse_labor_unit_basis(basis_raw)
        out.append({
            "laborUnitBasis": basis_letter,
            "laborUnitDivisor": basis_divisor,
            "laborUnitBasisRaw": str(basis_raw) if basis_raw is not None else None,
            "rowNumber": r,
            "itemId": item_id,
            "description": cell_or_none(ws, r, "B"),
            "category": cell_or_none(ws, r, "C"),
            "sector": cell_or_none(ws, r, "D"),
            "unit": cell_or_none(ws, r, "E"),
            "retailCost": as_number(cell_or_none(ws, r, "F")),
            "retailSource": cell_or_none(ws, r, "G"),
            "datePriced": str(cell_or_none(ws, r, "H")) if cell_or_none(ws, r, "H") else None,
            "tradeCost": as_number(cell_or_none(ws, r, "I")),
            # K (Cost Basis Used) is a FORMULA. It is deliberately NOT read as an
            # input -- the app resolves it from the supplier price table using the
            # same rule, so the resolution is testable rather than copied.
            "laborNormal": as_number(cell_or_none(ws, r, "L")),
            "laborDifficult": as_number(cell_or_none(ws, r, "M")),
            "laborVeryDifficult": as_number(cell_or_none(ws, r, "N")),
            "difficultyCurve": cell_or_none(ws, r, "O"),
            "necaUnitBasis": cell_or_none(ws, r, "P"),
            "necaPdfPage": cell_or_none(ws, r, "Q"),
            "laborStatus": cell_or_none(ws, r, "R"),
            "notes": cell_or_none(ws, r, "S"),
            "purchaseUnit": cell_or_none(ws, r, "T"),
            "purchasePackQty": as_number(cell_or_none(ws, r, "U")),
            "purchasePrice": as_number(cell_or_none(ws, r, "V")),
            "rowType": cell_or_none(ws, r, "Y"),
            "necArticle": cell_or_none(ws, r, "Z"),
        })
    return out, row_to_id


def parse_components(formula, pattern, row_to_id, assembly_id, which, problems):
    """Parse `<qty> * N(Atomics!K<row>)` terms out of an assembly's material formula."""
    if formula is None:
        return []
    rx = re.compile(pattern)
    comps = []
    for m in rx.finditer(str(formula)):
        row = int(m.group("row"))
        qty = float(m.group("qty"))
        item_id = row_to_id.get(row)
        if item_id is None:
            problems.append(
                f"{assembly_id}: {which} formula references Atomics row {row}, which holds no "
                f"Item ID. Row references break silently when rows are inserted or deleted."
            )
            continue
        comps.append({"itemId": item_id, "atomicRow": row, "quantity": qty})
    return comps


def extract_assemblies(ws, spec, sector, row_to_id, mapping, problems):
    ce = mapping["componentExtraction"]
    out = []
    for r in range(2, ws.max_row + 1):
        aid = cell_or_none(ws, r, "A")
        if aid is None:
            continue
        aid = str(aid).strip()

        cost_f = ws.cell(r, column_index_from_string("N")).value
        sell_f = ws.cell(r, column_index_from_string("O")).value
        labor_f = ws.cell(r, column_index_from_string("F")).value

        comps_cost = parse_components(cost_f, ce["materialCostFormulaPattern"], row_to_id, aid, "material-cost", problems)
        comps_sell = parse_components(sell_f, ce["materialSellFormulaPattern"], row_to_id, aid, "material-sell", problems)

        key_cost = sorted((c["itemId"], c["quantity"]) for c in comps_cost)
        key_sell = sorted((c["itemId"], c["quantity"]) for c in comps_sell)
        if key_cost != key_sell:
            problems.append(
                f"{aid}: material COST and material SELL formulas reference different "
                f"components. cost={key_cost} sell={key_sell}. This is the 2026-08-05 "
                f"'two columns, two sources' defect class and must be fixed in the workbook."
            )

        # knownStructuralFacts.assemblyLaborIsFrozenLiteral -- assert, do not assume.
        labor_refs_atomics = "Atomics!" in str(labor_f)

        out.append({
            "rowNumber": r,
            "assemblyId": aid,
            "name": cell_or_none(ws, r, "B"),
            "sectorColumn": cell_or_none(ws, r, "C"),
            "sectorTab": sector,
            "useCase": cell_or_none(ws, r, "D"),
            "componentProse": cell_or_none(ws, r, "E"),
            "totalLaborNormalFormula": str(labor_f) if labor_f is not None else None,
            "laborFormulaReferencesAtomics": labor_refs_atomics,
            "necCodeRefs": cell_or_none(ws, r, "H"),
            "notes": cell_or_none(ws, r, "I"),
            "status": cell_or_none(ws, r, "J"),
            "difficultySetting": cell_or_none(ws, r, "K"),
            "pricingFlags": cell_or_none(ws, r, "T"),
            "componentsTotalDeclared": as_number(cell_or_none(ws, r, "V")),
            "jobType": cell_or_none(ws, r, "AE"),
            "permitRequired": cell_or_none(ws, r, "AI"),
            "utilityStandbyRequired": cell_or_none(ws, r, "AJ"),
            "ceilingHeightBand": cell_or_none(ws, r, "AK"),
            "heightAccessAdderHours": as_number(cell_or_none(ws, r, "AL")),
            "sourcingChannel": cell_or_none(ws, r, "AM"),
            "necCategory": cell_or_none(ws, r, "AN"),
            "fieldDifficulty": cell_or_none(ws, r, "AO"),
            "components": comps_cost,
        })
    return out


def extract_suppliers(ws):
    out = []
    for r in range(2, ws.max_row + 1):
        sid = cell_or_none(ws, r, "A")
        if sid is None:
            continue
        out.append({
            "supplierId": str(sid).strip(),
            "name": cell_or_none(ws, r, "B"),
            "branch": cell_or_none(ws, r, "C"),
            "channel": cell_or_none(ws, r, "D"),
            "accountClass": cell_or_none(ws, r, "E"),
            "quotableRaw": cell_or_none(ws, r, "F"),
            "leadTime": cell_or_none(ws, r, "G"),
            "terms": cell_or_none(ws, r, "H"),
            "dateAdded": str(cell_or_none(ws, r, "I")) if cell_or_none(ws, r, "I") else None,
            "notes": cell_or_none(ws, r, "J"),
        })
    return out


def extract_supplier_prices(ws, duplicates):
    """Read the Supplier Prices tab in SHEET ORDER.

    Order is load-bearing. The workbook resolves cost with
    MATCH(key, 'Supplier Prices'!$L:$L, 0), and MATCH returns the FIRST match. If the
    same Item ID x Supplier appears twice, the row nearer the top is the one that
    prices every assembly — regardless of which is newer or better sourced. The
    importer relies on this list staying in sheet order so it can reproduce that
    behaviour exactly instead of letting a last-write-wins upsert quietly disagree
    with Excel.
    """
    out = []
    seen = {}
    for r in range(2, ws.max_row + 1):
        item_id = cell_or_none(ws, r, "A")
        supplier_id = cell_or_none(ws, r, "B")
        if item_id is None or supplier_id is None:
            continue
        key = f"{str(item_id).strip()}|{str(supplier_id).strip()}"
        if key in seen:
            duplicates.append({
                "key": key,
                "winningRow": seen[key]["row"],
                "winningPrice": seen[key]["price"],
                "winningDate": seen[key]["date"],
                "shadowedRow": r,
                "shadowedPrice": as_number(cell_or_none(ws, r, "C")),
                "shadowedDate": str(cell_or_none(ws, r, "G")) if cell_or_none(ws, r, "G") else None,
            })
        else:
            seen[key] = {
                "row": r,
                "price": as_number(cell_or_none(ws, r, "C")),
                "date": str(cell_or_none(ws, r, "G")) if cell_or_none(ws, r, "G") else None,
            }
        out.append({
            "rowNumber": r,
            "itemId": str(item_id).strip(),
            "supplierId": str(supplier_id).strip(),
            "priceAsPrinted": as_number(cell_or_none(ws, r, "C")),
            "pricedUom": cell_or_none(ws, r, "D"),
            "packQty": as_number(cell_or_none(ws, r, "E")),
            "datePriced": str(cell_or_none(ws, r, "G")) if cell_or_none(ws, r, "G") else None,
            "source": cell_or_none(ws, r, "H"),
            "availability": cell_or_none(ws, r, "I"),
            "accountClass": cell_or_none(ws, r, "J"),
            "quotableRaw": cell_or_none(ws, r, "K"),
            "confidence": cell_or_none(ws, r, "M"),
            "notes": cell_or_none(ws, r, "N"),
        })
    return out


def extract_nec_categories(ws):
    out = []
    for r in range(2, ws.max_row + 1):
        art = cell_or_none(ws, r, "A")
        if art is None:
            continue
        out.append({
            "article": str(art).strip(),
            "title": cell_or_none(ws, r, "B"),
            "onKyleList": cell_or_none(ws, r, "C"),
            "scopeRule": cell_or_none(ws, r, "D"),
        })
    return out


def extract_rate_config(ws, mapping):
    spec = mapping["rateConfig"]
    vcol = column_index_from_string(spec["valueColumn"])
    out = {}
    for key, cell in spec["cells"].items():
        raw = ws.cell(cell["row"], vcol).value
        if isinstance(raw, str) and raw.strip() == "":
            raw = None
        out[key] = {
            "row": cell["row"],
            "label": cell["label"],
            "raw": raw if not isinstance(raw, (int, float)) else raw,
            "number": as_number(raw),
            "text": str(raw) if raw is not None and not isinstance(raw, (int, float)) else None,
        }
    return out


# ─── Excel recalculation (the workbook's own computed truth) ─────────────────────

RECALC_SCRIPT = os.path.join(HERE, "recalc_workbook.ps1")


def recalc_with_excel(copy_path, mapping):
    """Recalculate the COPY in Excel and read back its computed values.

    Delegated to recalc_workbook.ps1 because this machine's Python has no COM bridge
    (no pywin32, no comtypes) while PowerShell reaches Excel directly. Installing a
    Python COM package would have been a change to the operator's machine to work
    around a capability that is already present -- so the pipeline uses what is there.

    Returns (data, None) on success or (None, reason) on failure. It NEVER returns a
    partial or synthesised result: without the workbook's own numbers the parity
    harness has no oracle, and a fabricated one is worse than no comparison at all.
    """
    if not os.path.isfile(RECALC_SCRIPT):
        return None, f"recalc script missing: {RECALC_SCRIPT}"

    out_json = os.path.join(os.path.dirname(copy_path), "recalc.json")
    exe = shutil.which("pwsh") or shutil.which("powershell")
    if exe is None:
        return None, "neither pwsh nor powershell found on PATH"

    cmd = [
        exe, "-NoProfile", "-NonInteractive", "-ExecutionPolicy", "Bypass",
        "-File", RECALC_SCRIPT,
        "-WorkbookCopy", copy_path,
        "-OutJson", out_json,
    ]
    try:
        proc = subprocess.run(cmd, capture_output=True, text=True, timeout=900)
    except subprocess.TimeoutExpired:
        return None, "Excel recalculation timed out after 900s"

    if proc.returncode != 0:
        detail = (proc.stderr or proc.stdout or "").strip().splitlines()
        tail = " | ".join(detail[-4:]) if detail else "no output"
        return None, f"recalc_workbook.ps1 exited {proc.returncode}: {tail}"

    if not os.path.isfile(out_json):
        return None, "recalc_workbook.ps1 reported success but wrote no JSON"

    try:
        with open(out_json, "r", encoding="utf-8-sig") as f:
            data = json.load(f)
    except (OSError, json.JSONDecodeError) as e:
        return None, f"could not read recalc JSON: {e}"

    for section in ("assemblies", "atomics", "supplierPrices"):
        if section not in data:
            return None, f"recalc JSON missing section {section!r}"
    return data, None


# ─── main ───────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(description="Extract price-book.xlsx to a JSON snapshot (read-only).")
    ap.add_argument("--workbook", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--no-recalc", action="store_true",
                    help="Skip the Excel recalculation. The snapshot then carries no "
                         "workbook-computed totals and the parity harness will refuse to pass.")
    args = ap.parse_args()

    wb_path = os.path.abspath(args.workbook)
    if not os.path.isfile(wb_path):
        print(f"FATAL: workbook not found: {wb_path}", file=sys.stderr)
        return 3

    with open(MAPPING_PATH, "r", encoding="utf-8") as f:
        mapping = json.load(f)

    hash_before = sha256(wb_path)

    tmpdir = tempfile.mkdtemp(prefix="pricebook-import-")
    copy_path = os.path.join(tmpdir, "price-book-readonly-copy.xlsx")
    shutil.copy2(wb_path, copy_path)

    try:
        wb = openpyxl.load_workbook(copy_path, data_only=False)

        drifts = Drifts()
        ws_atomics = validate_tab(wb, "atomics", mapping["tabs"]["atomics"], mapping, drifts)
        ws_res = validate_tab(wb, "assembliesResidential", mapping["tabs"]["assembliesResidential"], mapping, drifts)
        ws_com = validate_tab(wb, "assembliesCommercial", mapping["tabs"]["assembliesCommercial"], mapping, drifts)
        ws_sup = validate_tab(wb, "suppliers", mapping["tabs"]["suppliers"], mapping, drifts)
        ws_sp = validate_tab(wb, "supplierPrices", mapping["tabs"]["supplierPrices"], mapping, drifts)
        ws_nec = validate_tab(wb, "necCategoryMap", mapping["tabs"]["necCategoryMap"], mapping, drifts)
        ws_rate = validate_rate_config(wb, mapping, drifts)
        drifts.raise_if_any()

        problems = []
        atomics, row_to_id = extract_atomics(ws_atomics, mapping["tabs"]["atomics"])
        assemblies = (
            extract_assemblies(ws_res, mapping["tabs"]["assembliesResidential"],
                               mapping["tabs"]["assembliesResidential"]["sector"],
                               row_to_id, mapping, problems)
            + extract_assemblies(ws_com, mapping["tabs"]["assembliesCommercial"],
                                 mapping["tabs"]["assembliesCommercial"]["sector"],
                                 row_to_id, mapping, problems)
        )
        suppliers = extract_suppliers(ws_sup)
        duplicate_supplier_prices = []
        supplier_prices = extract_supplier_prices(ws_sp, duplicate_supplier_prices)
        nec_categories = extract_nec_categories(ws_nec)
        rate_config = extract_rate_config(ws_rate, mapping)

        # Assert the structural fact the pipeline is built on, rather than assuming it.
        labor_linked = [a["assemblyId"] for a in assemblies if a["laborFormulaReferencesAtomics"]]
        structural_notes = []
        if labor_linked:
            structural_notes.append(
                "CHANGED SINCE 2026-08-11: assembly labour formulas now reference the Atomics "
                f"tab on {len(labor_linked)} row(s): {labor_linked}. The mapping file records "
                "0 of 43. Re-read knownStructuralFacts.assemblyLaborIsFrozenLiteral before "
                "trusting any labour figure."
            )

        computed, recalc_error = (None, "skipped (--no-recalc)")
        if not args.no_recalc:
            computed, recalc_error = recalc_with_excel(copy_path, mapping)

        hash_after = sha256(wb_path)
        if hash_before != hash_after:
            print(
                "FATAL: the live workbook changed while it was being read.\n"
                f"  before: {hash_before}\n  after:  {hash_after}\n"
                "The snapshot is discarded rather than mixing two states. Re-run when the "
                "02:00/03:03 tasks or Excel are finished with the file.",
                file=sys.stderr,
            )
            return 4

        snapshot = {
            "generatedAt": datetime.now(timezone.utc).isoformat(),
            "mappingVersion": mapping["mappingVersion"],
            "workbookPath": wb_path,
            "workbookSha256": hash_before,
            "workbookSha256AfterRead": hash_after,
            "workbookMtime": datetime.fromtimestamp(os.path.getmtime(wb_path), timezone.utc).isoformat(),
            "readOnlyProof": "Live workbook was copied and never opened for write; SHA-256 identical before and after.",
            "counts": {
                "atomics": len(atomics),
                "assemblies": len(assemblies),
                "suppliers": len(suppliers),
                "supplierPrices": len(supplier_prices),
                "necCategories": len(nec_categories),
            },
            "rateConfig": rate_config,
            "atomics": atomics,
            "assemblies": assemblies,
            "suppliers": suppliers,
            "supplierPrices": supplier_prices,
            "necCategories": nec_categories,
            "workbookComputed": computed,
            "recalcError": recalc_error,
            "extractionProblems": problems,
            "structuralNotes": structural_notes,
            # Not fatal, but never silent: two rows for the same Item ID x Supplier
            # mean the workbook is quoting from whichever sits higher on the sheet.
            "duplicateSupplierPrices": duplicate_supplier_prices,
        }

        if duplicate_supplier_prices:
            print("\nDUPLICATE SUPPLIER PRICE ROWS — the workbook quotes the FIRST one:",
                  file=sys.stderr)
            for d in duplicate_supplier_prices:
                print(
                    f"  {d['key']}: row {d['winningRow']} (${d['winningPrice']}, {d['winningDate']}) "
                    f"WINS; row {d['shadowedRow']} (${d['shadowedPrice']}, {d['shadowedDate']}) "
                    f"is ignored by Excel's MATCH.",
                    file=sys.stderr,
                )

        if problems:
            print("EXTRACTION PROBLEMS (import will refuse to proceed):", file=sys.stderr)
            for p in problems:
                print(f"  - {p}", file=sys.stderr)

        os.makedirs(os.path.dirname(os.path.abspath(args.out)) or ".", exist_ok=True)
        with open(args.out, "w", encoding="utf-8") as f:
            json.dump(snapshot, f, indent=2, default=str)

        print(f"snapshot written: {args.out}")
        print(f"  atomics={len(atomics)} assemblies={len(assemblies)} "
              f"suppliers={len(suppliers)} supplierPrices={len(supplier_prices)}")
        print(f"  workbook sha256 unchanged: {hash_before[:16]}...")
        if computed is None:
            print(f"  WARNING: no workbook-computed values -- {recalc_error}")
        else:
            print(f"  workbook-computed assemblies: {len(computed['assemblies'])}")
        return 4 if problems else 0

    except ShapeDrift as e:
        print(str(e), file=sys.stderr)
        return 2
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


if __name__ == "__main__":
    sys.exit(main())
