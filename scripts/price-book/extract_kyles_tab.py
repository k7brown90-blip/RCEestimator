"""
Read `Atomics (Kyle's Copy)` and emit JSON on stdout. (P030)

READ-ONLY. This never opens Kyle's workbook for writing.

It reads CACHED VALUES (data_only=True) and, separately, the formula text, so it can tell the
difference between a cell that is genuinely blank and a cell holding a formula Excel has never
computed. The second case is reported as `uncomputed` and the importer stops on it — evaluating
a spreadsheet formula ourselves would be inventing a number Kyle did not.

Exit codes: 0 ok · 2 tab missing / shape wrong · 3 workbook unreadable.

Column layout (per the P030 mapping):
  A name · B/C/D labour Normal/Difficult/VeryDifficult · E company cost · F company price
  G/H/I RCE sell Normal/Difficult/VeryDifficult
A row with a name and no numbers anywhere in B–I is a SECTION HEADER.
"""

import json
import sys

try:
    import openpyxl
except ImportError:
    print("FATAL: openpyxl is required (pip install openpyxl)", file=sys.stderr)
    sys.exit(3)

NAME = 0
NUMERIC_COLS = [1, 2, 3, 4, 5, 6, 7, 8]  # B..I
# The cells whose absence-with-a-formula means "Excel has not computed this yet".
WATCH = {1: "LaborNormal", 5: "CompanyPrice", 6: "SellNormal", 7: "SellDifficult", 8: "SellVeryDifficult"}


def as_number(v):
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("$", "").replace(",", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def main() -> int:
    if len(sys.argv) < 3:
        print("usage: extract_kyles_tab.py <workbook> <tab>", file=sys.stderr)
        return 3
    path, tab = sys.argv[1], sys.argv[2]

    try:
        wbv = openpyxl.load_workbook(path, data_only=True, read_only=True)
        wbf = openpyxl.load_workbook(path, data_only=False, read_only=True)
    except Exception as exc:  # noqa: BLE001 - the caller only needs the message
        print(f"FATAL: cannot read workbook: {exc}", file=sys.stderr)
        return 3

    if tab not in wbv.sheetnames:
        print(f"FATAL: tab {tab!r} not in {wbv.sheetnames}", file=sys.stderr)
        return 2

    vrows = list(wbv[tab].iter_rows(values_only=True))
    frows = list(wbf[tab].iter_rows(values_only=True))
    wbv.close()
    wbf.close()

    rows, sections, uncomputed, unpriced = [], [], [], []
    current = None

    for i, vr in enumerate(vrows[1:], start=2):  # row 1 is the header
        fr = frows[i - 1] if i - 1 < len(frows) else ()
        raw_name = vr[NAME] if len(vr) > NAME else None
        name = str(raw_name).strip() if raw_name not in (None, "") else ""
        if not name:
            continue

        cells = [as_number(vr[c]) if len(vr) > c else None for c in NUMERIC_COLS]

        if all(c is None for c in cells):
            # Guard 1: a row whose formulas have not been computed is not a section header.
            pending = [
                f"row {i} {name!r} {label} holds a formula with no value"
                for col, label in WATCH.items()
                if len(fr) > col and isinstance(fr[col], str) and fr[col].startswith("=")
            ]
            if pending:
                uncomputed.extend(pending)
                continue

            # Guard 2: AN UNPRICED ITEM IS NOT A SECTION.
            #
            # "a name with no numbers" was a fine rule until Kyle added a LIGHTNING PROTECTION
            # section on 2026-08-18 with three product names under it and no prices yet. Those
            # three read as section headers, silently inflating the section count and vanishing
            # from the catalog — the worst outcome, because nothing said so.
            #
            # His section headers are short category labels and none of the 34 contains a comma;
            # every product name does ("Ridge Mount Base, Copper, 12-inch Strap..."). So a comma
            # means this is an item that has not been priced, and it gets REPORTED rather than
            # quietly reclassified.
            if "," in name:
                unpriced.append({"row": i, "name": name, "section": current})
                continue

            sections.append(name)
            current = name
            continue

        for col, label in WATCH.items():
            v = vr[col] if len(vr) > col else None
            f = fr[col] if len(fr) > col else None
            if v is None and isinstance(f, str) and f.startswith("="):
                uncomputed.append(f"row {i} {name!r} {label} holds a formula with no value")

        # The SELL formulas (G/H/I) travel with the row so parity can be asserted against the
        # formula the cell actually contains rather than against an assumption about the sheet.
        sell_formulas = [
            fr[c] if len(fr) > c and isinstance(fr[c], str) and fr[c].startswith("=") else None
            for c in (6, 7, 8)
        ]
        rows.append({
            "row": i, "name": name, "section": current, "cells": cells,
            "sellFormulas": sell_formulas,
        })

    json.dump(
        {"rows": rows, "sections": sections, "uncomputed": uncomputed, "unpriced": unpriced},
        sys.stdout,
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
